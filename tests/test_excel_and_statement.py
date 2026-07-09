"""v1.1: Excel exporter read-back and the contract full-picture statement."""

import datetime

from openpyxl import load_workbook

from govcon.models.billing import ScheduleType
from govcon.services.contract_statement import contract_statement
from govcon.services.export_excel import render_schedule_xlsx
from govcon.services.ice_schedules import BANNER, generate_schedule
from govcon.services.period_close import close_period
from tests.fixtures.synthetic_data import seed_all

D = datetime.date


def test_xlsx_roundtrip(session, tmp_path):
    data = seed_all(session)
    close_period(session, data.period_open, closed_by="xlsx-test")
    row = generate_schedule(session, 2026, ScheduleType.G)
    path = tmp_path / "schedule_g.xlsx"
    render_schedule_xlsx(row, str(path))
    wb = load_workbook(str(path))
    ws = wb.active
    assert ws["A1"].value == BANNER
    assert ws["A1"].font.bold is True
    cells = [c.value for r in ws.iter_rows() for c in r if c.value is not None]
    assert "reconciliation by period" in cells  # section title
    assert "period_id" in cells  # table header
    assert "2026-06" in cells  # a data value


def test_contract_statement_composes_full_picture(session):
    data = seed_all(session)
    contract = data.contracts["pre_ndaa"]
    statement = contract_statement(session, contract)
    assert statement["banner"] == BANNER
    assert statement["identity"]["contract_value"] == "12000000.00"
    # Immutable threshold regime rides along:
    assert statement["threshold_regime_immutable"]["tina_threshold_snapshot"] == "2500000.00"
    # CAS: $12M pre-NDAA = modified coverage.
    assert statement["cas_determination"]["tier"] == "modified"
    # Ledger composition: direct labor JCL and the unallowable split.
    assert statement["direct_costs_by_element"]["labor"] == "1250.00"
    gl = statement["gl_totals_by_cost_type"]
    assert gl["direct"] == "1250.00"
    assert gl["unallowable"] == "300.00"  # entertainment fixture — never billable
    assert statement["billing"]["billed_cumulative"] == "0.00"
    assert statement["audit_trail_rows_for_contract"] >= 1

"""Excel exporter (the committed next exporter per the 2026-07-08
export-format decision) — the same canonical-JSON walk as the markdown
exporter, rendered to a workbook: banner bold in A1, sections as stacked
blocks, list-of-dicts as header row + data rows.
"""

from __future__ import annotations

import json

from openpyxl import Workbook
from openpyxl.styles import Font

from govcon.models import ICESchedule
from govcon.services.export import _render_value  # shared scalar formatting
from govcon.services.ice_schedules import BANNER


def render_schedule_xlsx(schedule: ICESchedule, path: str) -> None:
    content = dict(json.loads(schedule.content))
    banner = content.pop("banner", BANNER)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Schedule {schedule.schedule_type.value} FY{schedule.fiscal_year}"

    ws["A1"] = banner
    ws["A1"].font = Font(bold=True, color="FFCC0000")
    ws.append([])
    ws.append([
        f"ICE Schedule {schedule.schedule_type.value} — FY{schedule.fiscal_year} "
        f"({schedule.reconciliation_status.value}, generated "
        f"{schedule.generated_date.isoformat()})"
    ])
    ws.append([])

    for key, value in content.items():
        if key in ("fiscal_year", "schedule"):
            continue
        title_cell_row = ws.max_row + 1
        ws.append([key.replace("_", " ")])
        ws.cell(row=title_cell_row, column=1).font = Font(bold=True)
        if isinstance(value, list) and value and all(isinstance(r, dict) for r in value):
            headers = list(value[0].keys())
            ws.append(headers)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
            for row in value:
                ws.append([_render_value(row.get(h)) for h in headers])
        elif isinstance(value, dict):
            for k, v in value.items():
                ws.append([k.replace("_", " "), _render_value(v)])
        else:
            ws.append([_render_value(value)])
        ws.append([])

    wb.save(path)
